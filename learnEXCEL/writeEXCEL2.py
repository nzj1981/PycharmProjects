# !/usr/bin/env python3.6
# coding=utf-8
import openpyxl
from openpyxl.utils import get_column_letter

# Write values to cell
wb = openpyxl.Workbook()

ws = wb.active

ws['A1'] = 'Hello Python'
print(ws['A1'].value)

# Create a new sheet and append values for cell
ws1 = wb.create_sheet('Range names')

for row in range(1, 41):
    ws1.append(range(17))
# Create a new 'List' of sheet and append values of list for cell
ws2 = wb.create_sheet('List')
rows = [
    ['Number', 'Batch 1', 'Batch 2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 60, 55],
    [6, 70, 10],
    [7, 80, 20]
]

for row in rows:
    ws2.append(row)
# Create a new 'Data' of sheet and add a value for the specified cell
ws3 = wb.create_sheet(title='Data', index=2)
for row in range(5, 30):
    for col in range(15, 54):
        ws3.cell(column=col, row=row, value=get_column_letter(col))
print(ws3['AA10'])

# Save alter.xlsx
wb.save(filename='alter.xlsx')