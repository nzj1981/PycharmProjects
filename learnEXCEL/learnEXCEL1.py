#!/usr/bin/env python3.6
# coding=utf-8
import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
# Getting sheets from the workbook
print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

# create a sheet page
mySheet = wb.create_sheet('mySheet')
print(wb.sheetnames)

# output a sheet page
sheet1 = wb['样例工作表']
print(sheet1)
sheet2 = wb.get_sheet_by_name('mySheet')
print(sheet2)

# Getting cells from the sheets
ws = wb.active
print(ws)
print(ws['a4'])
print(ws['A4'].value)

c = ws['B4']
print('Row {}, Column {} is {}'.format(c.row, c.column, c.value))
print('Cell {} is {}'.format(c.coordinate, c.value))

print(ws.cell(row=4, column=2))
print(ws.cell(row=4, column=2).value)

for i in range(4, 10, 2):
    print(i, ws.cell(row=i, column=2).value)

# Getting rows and columns from the sheets
colC = ws['C']
print(colC)
print(colC[4].value)
row4 = ws['4']
print(row4)

col_range = ws['B:C']
row_range = ws[4:6]

for col in col_range:
    for cell in col:
        print(cell.value)
print("================*******************+++++++++++++++++++++")
for row in row_range:
    for cell in row:
        print(cell.value)
print("*******************************************************")
for row in ws.iter_rows(min_row=8, max_row=10, max_col=2):
    for cell in row:
        print(cell.value)
for col in ws.iter_rows(min_col=2, max_col=4, max_row=4):
    for cell in col:
        print(cell.value)

# ws.rows type is tuple
print(tuple(ws.rows))
print("*******************************************************")
# Getting cell value
cell_range = ws['B4:d6']

for rowOfCellOjbects in cell_range:
    for cellObj in rowOfCellOjbects:
        print(cellObj.coordinate, cellObj.value)
    print('____________End of Row_________________')
# output max_row and max_column
print('{} * {}'.format(ws.max_row, ws.max_column))

# Getting letter to index or index to letter
from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(2), get_column_letter(47), get_column_letter(256))
print(column_index_from_string('AB'), column_index_from_string('cd'), column_index_from_string("bd"))
