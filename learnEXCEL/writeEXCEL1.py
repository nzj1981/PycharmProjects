#!/usr/bin/env python3.6
# coding=utf-8

import openpyxl

wb = openpyxl.Workbook()

ws = wb.active
# change the name of the sheet
print(ws.title)
ws.title = 'Happy2017'
print(wb.get_sheet_names())
wb.save('example_write.xlsx')

# Save an EXCEL
# wb = openpyxl.load_workbook('example.xlsx')
#
# wb.save('alter.xlsx')

# Add one or more sheets for EXCEL

wb.create_sheet(title='First Sheet', index=0)
wb.create_sheet(index=1, title='Middle Sheet')
print(wb.get_sheet_names())

# copy one or more sheets for EXCEl
ws1 = wb.copy_worksheet(ws)
ws1.title = 'Middle Sheet two'
ws2 = wb.copy_worksheet(wb.get_sheet_by_name('First Sheet'))

# remove one or more sheets for EXCEL
wb.remove(wb.get_sheet_by_name('Middle Sheet'))
print(wb.get_sheet_names())

# Save new example_write.xlsx
wb.save('example_write.xlsx')
