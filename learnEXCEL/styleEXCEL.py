# !/usr/bin/env python3.6
# coding=utf-8
import datetime

import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import colors

# open new excel
wb = openpyxl.Workbook()
ws = wb.active
# Font
ws.title = 'Font'
# default 11pt,宋体
italic24Font = Font(size=24, italic=True)
ws['B3'].font = italic24Font
ws['B3'] = '24pt Italic'
ws['B2'].font = italic24Font
ws.cell(row=2, column=2).value = '宋体字'

boldRedFont = Font(size=12, name='Times New Roman', bold=True, color=colors.RED)
ws['A1'].font = boldRedFont
ws['A1'] = 'Bold Red Times New Roman'

# Formulas
ws = wb.copy_worksheet(ws)
ws.title = 'Formulas'
ws['A4'] = 200
ws['A5'] = 300
ws['A6'] = '=SUM(A4:A5)'
ws['A7'] = datetime.datetime.now()

# Using number formats
ws = wb.copy_worksheet(wb.get_sheet_by_name('Font'))
ws.title = 'Number Formats'
# Set date using a Python datetime
ws['A4'] = datetime.datetime(2017, 12, 21)
# format 'yyyy-mm-dd h:mm:ss'
ws['A4'].number_format
# you can enable type inference on a case-by-case basis
wb.guess_types = True
# set percentage using a string followed by the percent sign
ws['B4'] = '3.14%'
wb.guess_types = False
ws['B4'].value
ws['B4'].number_format

# Setting row height and column width
ws = wb.create_sheet('dimensions')
ws['A1'] = 'Tall row'
ws.row_dimensions[1].height = 70
ws['B2'] = 'Wide column'
ws.column_dimensions['B'].width = 20

# Merging cells
ws = wb.create_sheet('merged')
ws.merge_cells('A1:D3')
ws['A1'] = 'Twelve cells merged together'
# or equivalently
ws.merge_cells(start_row=4, start_column=5, end_row=7, end_column=8)
ws['E4'] = 'equivalently'

# Unmerging cells
ws = wb.copy_worksheet(wb.get_sheet_by_name('merged'))
ws.title = 'unmerged'
ws.unmerge_cells('A1:D3')
# or equivalently
ws.unmerge_cells(start_row=4, start_column=5, end_row=7, end_column=8)

# Save excel

wb.save('style.xlsx')
