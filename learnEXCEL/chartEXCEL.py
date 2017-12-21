# !/usr/bin/env python3.6
# coding=utf-8

import openpyxl
from openpyxl.chart import (
    Reference,
    Series,
    PieChart,
    BarChart,
    BubbleChart
)
from openpyxl.chart.series import DataPoint

# open new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'picChart'
# Pie chart

data = [
    ['Pie', 'Sold'],
    ['Apple', 50],
    ['Cherry', 30],
    ['Pumpkin', 10],
    ['Chocolate', 40]
]

for row in data:
    ws.append(row)

pie = PieChart()
data = Reference(ws, min_col=2, min_row=2, max_row=5)
labels = Reference(ws, min_col=1, min_row=2, max_row=5)
# "set_categories" must be behind "add_data", otherwise the legend shows list of numbers
pie.add_data(data)
pie.set_categories(labels)
pie.title = 'Pies sold by category'

# Cut the first slice out of the pie
slice1 = DataPoint(idx=0, explosion=10)
pie.series[0].data_points = [slice1]

ws.add_chart(pie, 'A15')

# Bar chart
rows = [
    ('Number', 'Batch 1', 'Batch 2', 'Batch 3'),
    (2, 10, 30, 100),
    (3, 40, 60, 50),
    (4, 50, 70, 80),
    (5, 20, 10, 40),
    (6, 10, 40, 10),
    (7, 50, 30, 20),
]
ws = wb.create_sheet('Bar Chart')
for row in rows:
    ws.append(row)

chart1 = BarChart()
chart1.type = 'col'
chart1.style = 10
chart1.title = 'Bar Chart'
chart1.y_axis.title = 'Sample length(mm)'
chart1.x_axis.title = 'Test number'

data = Reference(ws, min_col=2, min_row=2, max_row=7, max_col=4)
cats = Reference(ws, min_col=1, min_row=2, max_row=7)
chart1.add_data(data)
chart1.set_categories(cats)
ws.add_chart(chart1, 'A10')

# Using "deepcopy"
from copy import deepcopy

chart2 = deepcopy(chart1)
chart2.type = "bar"
chart2.style = 10
chart2.title = 'Horizontal Bar Chart'
ws.add_chart(chart2, "J10")

# Bubble Chart
rows = [
    ("Number of Products", "Sales in USD", "Market share"),
    (14, 12200, 15),
    (20, 60000, 33),
    (18, 24400, 10),
    (22, 32000, 42),
    (),
    (12, 8200, 18),
    (15, 50000, 30),
    (19, 22400, 15),
    (25, 25000, 50),
]

ws = wb.create_sheet('Bubble Chart')
for row in rows:
    ws.append(row)
chart = BubbleChart()
# use a preset style
chart.style = 18
# add the first series of data
xvalues = Reference(ws, min_col=1, min_row=2, max_row=5)
yvalues = Reference(ws, min_col=2, min_row=2, max_row=5)
size = Reference(ws, min_col=3, min_row=2, max_row=5)
series = Series(values=yvalues, xvalues=xvalues, zvalues=size, title='2015')
chart.series.append(series)

# add the second
xvalues = Reference(ws, min_col=1, min_row=7, max_row=10)
yvalues = Reference(ws, min_col=2, min_row=7, max_row=10)
size = Reference(ws, min_col=3, min_row=7, max_row=10)
series = Series(values=yvalues, xvalues=xvalues, zvalues=size, title='2016')
chart.series.append(series)
# place the chart starting in cell E1
ws.add_chart(chart, 'E1')


# Save pie chart

wb.save('chart.xlsx')
