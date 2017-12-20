#!/usr/bin/env python3.6
# coding=utf-8

import openpyxl,pprint

'''
{'河南省':{'郑州市':{'coding':410108, 'pop':1999},
          '焦作市':{'coding':410882, 'pop':15992},
          '三门峡市':{'coding':411221, 'pop':115000} 
    }, 
'河北省':{ 

    }

}
countryData['河南省']['郑州市']['coding']

'''

# Read the sheet1 data
print('Opening workbook')
wb = openpyxl.load_workbook('example2.xlsx')
# wb = openpyxl.load_workbook('censuspopdata.xlsx')
ws = wb.active
countryData = {}

# Fill in countryData with each city's pop and coding
for row in range(2, ws.max_row+1):
    # Each row in the sheet1 has data
    state = ws['B' + str(row)].value
    country = ws['C' + str(row)].value
    pop = ws['D' + str(row)].value
    # make sure the key state exists
    countryData.setdefault(state, {})
    # make sure the key for country in state exists
    countryData[state].setdefault(country, {'coding': 0, 'pop': 0})
    # Each row represents one census tract, so increment by one
    countryData[state][country]['coding'] += 1
    # Increase the country pop by the pop in this census tract
    countryData[state][country]['pop'] += int(pop)
# Open a new text file and write the contents for countryData to it
print('Writing results...')
resultFile = open('census2017.py', 'w', encoding='utf8')
resultFile.writelines('#!/usr/bin/env python3.6\n')
resultFile.writelines('# coding=utf-8\n\n')
resultFile.write('allData = ' + pprint.pformat(countryData))
resultFile.close()